"""Phrase whitelist service for spec 011 Layer D matching.

Exposes normalize_phrase plus CRUD operations for pair-level and phrase-level
whitelist management.
"""

import csv
import re
import sqlite3
import unicodedata
from datetime import UTC, datetime
from pathlib import Path

from tube_scout.models.reuse_v2 import (
    MatchSpan,
    WhitelistPairEntry,
    WhitelistPhraseEntry,
    WhitelistView,
)

# Korean and ASCII punctuation to strip (research.md R-7)
_PUNCT_PATTERN = re.compile(
    r"[。、，．・「」『』""''‥…—–·〈〉《》【】｢｣〔〕"
    r',.!?;:()[\]{}\-—–…"\'`~@#$%^&*+=|\\/<>]'
)


def normalize_phrase(text: str) -> str:
    """Apply 5-step normalization for exact-equality phrase comparison.

    Steps (research.md R-7):
      1. Unicode NFKC normalization (full-width → half-width, glyph unification).
      2. casefold (lowercases English; no effect on Korean/CJK).
      3. Punctuation strip — Korean punct + ASCII punct set removed.
      4. Whitespace collapse — any run of whitespace (incl. tab, newline,
         NBSP, IDEOGRAPHIC SPACE) → single ASCII space.
      5. Strip leading/trailing whitespace.

    Args:
        text: Raw phrase or caption segment text to normalize.

    Returns:
        Normalized string suitable for exact-equality comparison.
        Empty string is returned for inputs that are entirely whitespace
        or punctuation.
    """
    # Step 1: NFKC
    normalized = unicodedata.normalize("NFKC", text)
    # Step 2: casefold
    normalized = normalized.casefold()
    # Step 3: strip punctuation
    normalized = _PUNCT_PATTERN.sub(" ", normalized)
    # Step 4: collapse all whitespace variants to single space
    normalized = re.sub(r"\s+", " ", normalized)
    # Step 5: trim
    return normalized.strip()


def add_pair_whitelist(
    source_video_id: str,
    target_video_id: str,
    reason: str,
    db_path: Path,
    registered_by: str,
) -> int:
    """Mark a video pair as FALSE_POSITIVE in comparison_results.

    Args:
        source_video_id: Source video identifier.
        target_video_id: Target video identifier.
        reason: Admin explanation for the exclusion.
        db_path: SQLite content_reuse.db path.
        registered_by: Admin identifier.

    Returns:
        The comparison_results row id that was updated.

    Raises:
        TypeError: If db_path is not a Path.
        ValueError: If the pair is not found in comparison_results.
    """
    if not isinstance(db_path, Path):
        raise TypeError(f"db_path must be a Path, got {type(db_path).__name__}")

    conn = sqlite3.connect(str(db_path))
    try:
        row = conn.execute(
            "SELECT id FROM comparison_results "
            "WHERE source_video_id = ? AND target_video_id = ?",
            (source_video_id, target_video_id),
        ).fetchone()
        if row is None:
            raise ValueError(
                f"pair not found: source={source_video_id!r} target={target_video_id!r}"
            )
        comp_id: int = row[0]
        conn.execute(
            "UPDATE comparison_results SET review_status = 'FALSE_POSITIVE', "
            "reviewed_at = ?, reviewed_by = ? WHERE id = ?",
            (datetime.now(UTC).isoformat(), registered_by, comp_id),
        )
        conn.commit()
    finally:
        conn.close()

    return comp_id


def add_phrase_whitelist(
    professor_id: str,
    phrase_raw: str,
    reason: str,
    db_path: Path,
    registered_by: str,
) -> WhitelistPhraseEntry:
    """Insert a normalized phrase into phrase_whitelist for the given professor.

    Args:
        professor_id: Professor pool identifier.
        phrase_raw: Original phrase text (stored for display/audit).
        reason: Admin explanation for the whitelist entry.
        db_path: SQLite content_reuse.db path.
        registered_by: Admin identifier.

    Returns:
        WhitelistPhraseEntry representing the inserted row.

    Raises:
        TypeError: If db_path is not a Path.
        ValueError: If this (professor_id, phrase_normalized) already exists.
    """
    if not isinstance(db_path, Path):
        raise TypeError(f"db_path must be a Path, got {type(db_path).__name__}")

    phrase_normalized = normalize_phrase(phrase_raw)
    now = datetime.now(UTC).isoformat()

    conn = sqlite3.connect(str(db_path))
    try:
        try:
            conn.execute(
                "INSERT INTO phrase_whitelist "
                "(professor_id, phrase_normalized, phrase_raw, reason, registered_by, registered_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (professor_id, phrase_normalized, phrase_raw, reason, registered_by, now),
            )
            conn.commit()
        except sqlite3.IntegrityError as exc:
            raise ValueError(
                f"already whitelisted: professor_id={professor_id!r} phrase={phrase_raw!r}"
            ) from exc
    finally:
        conn.close()

    return WhitelistPhraseEntry(
        professor_id=professor_id,
        phrase_normalized=phrase_normalized,
        phrase_raw=phrase_raw,
        reason=reason,
        admin=registered_by,
        registered_at=datetime.fromisoformat(now),
    )


def list_whitelist(
    db_path: Path,
    professor_id: str | None = None,
    kind: str | None = None,
) -> WhitelistView:
    """Return whitelist entries, optionally filtered by professor and/or kind.

    Args:
        db_path: SQLite content_reuse.db path.
        professor_id: If provided, filter phrase entries by this professor.
        kind: 'pair', 'phrase', or None (both).

    Returns:
        WhitelistView with pair_entries and phrase_entries populated.

    Raises:
        TypeError: If db_path is not a Path.
        ValueError: If kind is not 'pair', 'phrase', or None.
    """
    if not isinstance(db_path, Path):
        raise TypeError(f"db_path must be a Path, got {type(db_path).__name__}")
    if kind not in (None, "pair", "phrase"):
        raise ValueError(f"kind must be 'pair', 'phrase', or None, got {kind!r}")

    conn = sqlite3.connect(str(db_path))
    try:
        pair_entries: list[WhitelistPairEntry] = []
        phrase_entries: list[WhitelistPhraseEntry] = []

        if kind in (None, "pair"):
            rows = conn.execute(
                "SELECT source_video_id, target_video_id, review_status, "
                "reviewed_at, reviewed_by "
                "FROM comparison_results WHERE review_status = 'FALSE_POSITIVE'"
            ).fetchall()
            for row in rows:
                pair_entries.append(
                    WhitelistPairEntry(
                        source_video_id=row[0],
                        target_video_id=row[1],
                        reason="FALSE_POSITIVE",
                        admin=row[4] or "unknown",
                        registered_at=datetime.fromisoformat(row[3])
                        if row[3]
                        else datetime.now(UTC),
                    )
                )

        if kind in (None, "phrase"):
            if professor_id is not None:
                rows = conn.execute(
                    "SELECT professor_id, phrase_normalized, phrase_raw, "
                    "reason, registered_by, registered_at "
                    "FROM phrase_whitelist WHERE professor_id = ?",
                    (professor_id,),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT professor_id, phrase_normalized, phrase_raw, "
                    "reason, registered_by, registered_at "
                    "FROM phrase_whitelist"
                ).fetchall()
            for row in rows:
                phrase_entries.append(
                    WhitelistPhraseEntry(
                        professor_id=row[0],
                        phrase_normalized=row[1],
                        phrase_raw=row[2],
                        reason=row[3],
                        admin=row[4] or "unknown",
                        registered_at=datetime.fromisoformat(row[5]),
                    )
                )
    finally:
        conn.close()

    return WhitelistView(pair_entries=pair_entries, phrase_entries=phrase_entries)


def export_whitelist(
    db_path: Path,
    fmt: str,
    output_path: Path,
) -> Path:
    """Export whitelist entries to a file in the requested format.

    Args:
        db_path: SQLite content_reuse.db path.
        fmt: Output format — 'csv', 'xlsx', or 'markdown'.
        output_path: Destination file path.

    Returns:
        output_path after writing.

    Raises:
        TypeError: If db_path or output_path is not a Path.
        ValueError: If fmt is not a supported format.
    """
    if not isinstance(db_path, Path):
        raise TypeError(f"db_path must be a Path, got {type(db_path).__name__}")
    if not isinstance(output_path, Path):
        raise TypeError(f"output_path must be a Path, got {type(output_path).__name__}")
    if fmt not in ("csv", "xlsx", "markdown"):
        raise ValueError(f"Unsupported format: {fmt!r}. Use 'csv', 'xlsx', or 'markdown'.")

    view = list_whitelist(db_path)

    pair_rows = [
        {
            "kind": "pair",
            "source_video_id": e.source_video_id,
            "target_video_id": e.target_video_id,
            "professor_id": e.professor_id or "",
            "phrase_raw": "",
            "reason": e.reason,
            "admin": e.admin,
            "registered_at": e.registered_at.isoformat(),
        }
        for e in view.pair_entries
    ]
    phrase_rows = [
        {
            "kind": "phrase",
            "source_video_id": "",
            "target_video_id": "",
            "professor_id": e.professor_id,
            "phrase_raw": e.phrase_raw,
            "reason": e.reason,
            "admin": e.admin,
            "registered_at": e.registered_at.isoformat(),
        }
        for e in view.phrase_entries
    ]

    fieldnames = [
        "kind", "source_video_id", "target_video_id",
        "professor_id", "phrase_raw", "reason", "admin", "registered_at",
    ]

    if fmt == "csv":
        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(pair_rows + phrase_rows)

    elif fmt == "xlsx":
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        pair_ws = wb.create_sheet("Pair")
        pair_ws.append(fieldnames)
        for r in pair_rows:
            pair_ws.append([r[f] for f in fieldnames])

        phrase_ws = wb.create_sheet("Phrase")
        phrase_ws.append(fieldnames)
        for r in phrase_rows:
            phrase_ws.append([r[f] for f in fieldnames])

        wb.save(str(output_path))

    elif fmt == "markdown":
        header = "| " + " | ".join(fieldnames) + " |"
        separator = "| " + " | ".join(["---"] * len(fieldnames)) + " |"
        lines = [header, separator]
        for row in pair_rows + phrase_rows:
            lines.append("| " + " | ".join(str(row[f]) for f in fieldnames) + " |")
        output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    return output_path


def remove_whitelist(
    db_path: Path,
    kind: str,
    entry_id: int,
) -> bool:
    """Remove a whitelist entry by kind and id.

    For kind='phrase': deletes the row from phrase_whitelist.
    For kind='pair': resets comparison_results.review_status to UNREVIEWED.

    Args:
        db_path: SQLite content_reuse.db path.
        kind: 'pair' or 'phrase'.
        entry_id: Row id to remove.

    Returns:
        True if an entry was removed/reset.

    Raises:
        TypeError: If db_path is not a Path.
        ValueError: If kind is not 'pair' or 'phrase'.
    """
    if not isinstance(db_path, Path):
        raise TypeError(f"db_path must be a Path, got {type(db_path).__name__}")
    if kind not in ("pair", "phrase"):
        raise ValueError(f"kind must be 'pair' or 'phrase', got {kind!r}")

    conn = sqlite3.connect(str(db_path))
    try:
        if kind == "phrase":
            cur = conn.execute(
                "DELETE FROM phrase_whitelist WHERE id = ?", (entry_id,)
            )
            conn.commit()
            return cur.rowcount > 0
        else:
            cur = conn.execute(
                "UPDATE comparison_results SET review_status = 'UNREVIEWED', "
                "reviewed_at = NULL, reviewed_by = NULL WHERE id = ?",
                (entry_id,),
            )
            conn.commit()
            return cur.rowcount > 0
    finally:
        conn.close()


def subtract_phrase_whitelist(
    professor_id: str,
    spans: list[MatchSpan],
    db_path: Path,
) -> tuple[list[MatchSpan], int]:
    """Remove spans whose matched_text_sample matches a whitelisted phrase.

    Args:
        professor_id: Professor pool identifier for whitelist lookup.
        spans: MatchSpan list to filter.
        db_path: SQLite content_reuse.db path.

    Returns:
        Tuple of (remaining spans, number of spans removed).

    Raises:
        TypeError: If db_path is not a Path.
    """
    if not isinstance(db_path, Path):
        raise TypeError(f"db_path must be a Path, got {type(db_path).__name__}")

    conn = sqlite3.connect(str(db_path))
    try:
        rows = conn.execute(
            "SELECT phrase_normalized FROM phrase_whitelist WHERE professor_id = ?",
            (professor_id,),
        ).fetchall()
    finally:
        conn.close()

    whitelist_norms: set[str] = {row[0] for row in rows}
    remaining: list[MatchSpan] = []
    removed_count = 0

    for span in spans:
        norm = normalize_phrase(span.matched_text_sample)
        if norm in whitelist_norms:
            removed_count += 1
        else:
            remaining.append(span)

    return remaining, removed_count
