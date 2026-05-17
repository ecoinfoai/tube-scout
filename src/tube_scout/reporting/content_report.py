"""Content quality report generator.

Generates HTML, Excel, and JSON reports from comparison results
and quality check data for administrator review.

Also exposes generate_v2_report() for spec 011 nC2 4-pattern reports (US5).
"""

import json
import logging
import sqlite3
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal

logger = logging.getLogger(__name__)

import openpyxl
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Font, PatternFill


def _sanitize_cell(value: Any) -> Any:
    """Sanitize a cell value to prevent Excel formula injection.

    Args:
        value: Cell value to sanitize.

    Returns:
        Sanitized value safe for Excel.
    """
    if not isinstance(value, str):
        return value
    if value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _build_summary(comparisons: list[dict[str, Any]]) -> dict[str, Any]:
    """Build summary statistics from comparison results.

    Args:
        comparisons: List of comparison result dicts.

    Returns:
        Summary dict with counts and statistics.
    """
    grade_counts = Counter(c.get("grade", "unknown") for c in comparisons)
    review_counts = Counter(c.get("review_status", "UNREVIEWED") for c in comparisons)

    # Per-professor suspicion rates
    professor_scores: dict[str, list[float]] = {}
    for c in comparisons:
        prof = c.get("professor", "Unknown")
        score = c.get("suspicion_score", 0.0) or 0.0
        professor_scores.setdefault(prof, []).append(score)

    professor_rates = {
        prof: round(sum(scores) / len(scores), 2)
        for prof, scores in professor_scores.items()
    }

    return {
        "total_comparisons": len(comparisons),
        "grade_counts": {
            "critical": grade_counts.get("critical", 0),
            "high": grade_counts.get("high", 0),
            "moderate": grade_counts.get("moderate", 0),
            "normal": grade_counts.get("normal", 0),
        },
        "review_counts": dict(review_counts),
        "professor_avg_suspicion": professor_rates,
    }


class ContentReportGenerator:
    """Generator for content quality reports in multiple formats."""

    def generate_json(
        self,
        comparisons: list[dict[str, Any]],
        quality_results: list[dict[str, Any]],
        output_path: Path,
    ) -> None:
        """Generate JSON report.

        Args:
            comparisons: Comparison result dicts.
            quality_results: Quality check result dicts.
            output_path: Output file path.
        """
        report = {
            "summary": _build_summary(comparisons),
            "comparisons": comparisons,
            "quality_results": quality_results,
        }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )

    def generate_html(
        self,
        comparisons: list[dict[str, Any]],
        quality_results: list[dict[str, Any]],
        output_path: Path,
    ) -> None:
        """Generate HTML report.

        Args:
            comparisons: Comparison result dicts.
            quality_results: Quality check result dicts.
            output_path: Output file path.
        """
        summary = _build_summary(comparisons)

        grade_colors = {
            "critical": "#dc3545",
            "high": "#fd7e14",
            "moderate": "#ffc107",
            "normal": "#28a745",
        }

        # Build comparison rows
        rows_html = ""
        sorted_comps = sorted(
            comparisons,
            key=lambda x: x.get("suspicion_score", 0),
            reverse=True,
        )
        for c in sorted_comps:
            grade = c.get("grade", "")
            color = grade_colors.get(grade, "#6c757d")
            prof = c.get("professor", "")
            course = c.get("course", "")
            ws = f"W{c.get('week', '?')}/S{c.get('session', '?')}"
            yrs = f"{c.get('year_from', '')}->{c.get('year_to', '')}"
            score = c.get("suspicion_score", 0)
            status = c.get("review_status", "")
            rows_html += f"""
            <tr>
                <td>{prof}</td>
                <td>{course}</td>
                <td>{ws}</td>
                <td>{yrs}</td>
                <td style="color: {color}; font-weight: bold;">{score:.1f}</td>
                <td style="color: {color};">{grade}</td>
                <td>{status}</td>
            </tr>"""

        # Build quality rows
        quality_html = ""
        for q in quality_results:
            quality_html += f"""
            <tr>
                <td>{q.get('video_id', '')}</td>
                <td>{'Pass' if q.get('q001_voice_present') else 'Fail'}</td>
                <td>{'Pass' if q.get('q002_min_duration') else 'Fail'}</td>
                <td>{q.get('q003_course_relevance', 'N/A')}</td>
                <td>{q.get('q004_silence_ratio', 'N/A')}</td>
                <td>{q.get('q005_speech_density', 'N/A')}</td>
                <td>{q.get('pass_count', 0)}/5</td>
            </tr>"""

        gc = summary["grade_counts"]
        cnt_crit = gc["critical"]
        cnt_high = gc["high"]
        cnt_mod = gc["moderate"]
        cnt_norm = gc["normal"]

        html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="utf-8">
    <title>Content Quality Report</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont,
                'Segoe UI', sans-serif;
            margin: 20px;
        }}
        h1 {{ color: #333; }}
        h2 {{ color: #555; margin-top: 30px; }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin-top: 10px;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        th {{ background-color: #f8f9fa; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        .summary {{
            display: flex; gap: 20px; margin: 20px 0;
        }}
        .summary-card {{
            padding: 15px;
            border-radius: 8px;
            min-width: 120px;
            text-align: center;
        }}
        .critical {{ background-color: #f8d7da; color: #721c24; }}
        .high {{ background-color: #fff3cd; color: #856404; }}
        .moderate {{ background-color: #d1ecf1; color: #0c5460; }}
        .normal {{ background-color: #d4edda; color: #155724; }}
    </style>
</head>
<body>
    <h1>Content Quality Report</h1>

    <h2>Summary</h2>
    <div class="summary">
        <div class="summary-card critical">
            <div style="font-size: 24px; font-weight: bold;">{cnt_crit}</div>
            <div>Critical</div>
        </div>
        <div class="summary-card high">
            <div style="font-size: 24px; font-weight: bold;">{cnt_high}</div>
            <div>High</div>
        </div>
        <div class="summary-card moderate">
            <div style="font-size: 24px; font-weight: bold;">{cnt_mod}</div>
            <div>Moderate</div>
        </div>
        <div class="summary-card normal">
            <div style="font-size: 24px; font-weight: bold;">{cnt_norm}</div>
            <div>Normal</div>
        </div>
    </div>

    <h2>Suspicion Results</h2>
    <table>
        <thead>
            <tr>
                <th>Professor</th>
                <th>Course</th>
                <th>Week/Session</th>
                <th>Years</th>
                <th>Score</th>
                <th>Grade</th>
                <th>Review Status</th>
            </tr>
        </thead>
        <tbody>{rows_html}
        </tbody>
    </table>

    <h2>Quality Checklist</h2>
    <table>
        <thead>
            <tr>
                <th>Video ID</th>
                <th>Q-001 Voice</th>
                <th>Q-002 Duration</th>
                <th>Q-003 Relevance</th>
                <th>Q-004 Silence</th>
                <th>Q-005 Density</th>
                <th>Pass</th>
            </tr>
        </thead>
        <tbody>{quality_html}
        </tbody>
    </table>
</body>
</html>"""

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(html, encoding="utf-8")

    def generate_xlsx(
        self,
        comparisons: list[dict[str, Any]],
        quality_results: list[dict[str, Any]],
        output_path: Path,
    ) -> None:
        """Generate Excel report with multiple sheets.

        Args:
            comparisons: Comparison result dicts.
            quality_results: Quality check result dicts.
            output_path: Output file path.
        """
        wb = openpyxl.Workbook()
        header_font = Font(bold=True)

        # Sheet 1: Suspicion Summary
        ws_susp = wb.active
        ws_susp.title = "Suspicion Summary"
        headers = [
            "ID", "Professor", "Course", "Week", "Session",
            "Year From", "Year To", "Hash Match", "Cosine Sim",
            "Change Rate", "New Terms", "Duration Diff",
            "Score", "Grade", "Review Status",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws_susp.cell(row=1, column=col, value=h)
            cell.font = header_font

        grade_fills = {
            "critical": PatternFill(
                start_color="DC3545", end_color="DC3545",
                fill_type="solid",
            ),
            "high": PatternFill(
                start_color="FD7E14", end_color="FD7E14",
                fill_type="solid",
            ),
            "moderate": PatternFill(
                start_color="FFC107", end_color="FFC107",
                fill_type="solid",
            ),
            "normal": PatternFill(
                start_color="28A745", end_color="28A745",
                fill_type="solid",
            ),
        }

        sorted_comps = sorted(
            comparisons,
            key=lambda x: x.get("suspicion_score", 0),
            reverse=True,
        )
        for row_idx, c in enumerate(sorted_comps, start=2):
            values = [
                c.get("id"),
                _sanitize_cell(c.get("professor", "")),
                _sanitize_cell(c.get("course", "")),
                c.get("week"),
                c.get("session"),
                c.get("year_from"),
                c.get("year_to"),
                bool(c.get("i1_hash_match")),
                c.get("i2_cosine_similarity"),
                c.get("i3_change_rate"),
                c.get("i4_new_term_count"),
                c.get("i5_duration_diff_seconds"),
                c.get("suspicion_score"),
                c.get("grade", ""),
                c.get("review_status", ""),
            ]
            for col, val in enumerate(values, 1):
                ws_susp.cell(row=row_idx, column=col, value=val)

            grade = c.get("grade", "")
            if grade in grade_fills:
                ws_susp.cell(row=row_idx, column=14).fill = grade_fills[grade]

        # Sheet 2: Quality Results
        ws_qual = wb.create_sheet("Quality Results")
        q_headers = [
            "Video ID", "Q-001 Voice", "Q-002 Duration",
            "Q-003 Relevance", "Q-004 Silence", "Q-005 Density", "Pass Count",
        ]
        for col, h in enumerate(q_headers, 1):
            cell = ws_qual.cell(row=1, column=col, value=h)
            cell.font = header_font

        for row_idx, q in enumerate(quality_results, start=2):
            values = [
                _sanitize_cell(q.get("video_id", "")),
                bool(q.get("q001_voice_present")),
                bool(q.get("q002_min_duration")),
                q.get("q003_course_relevance"),
                q.get("q004_silence_ratio"),
                q.get("q005_speech_density"),
                q.get("pass_count", 0),
            ]
            for col, val in enumerate(values, 1):
                ws_qual.cell(row=row_idx, column=col, value=val)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        wb.close()


# ---------------------------------------------------------------------------
# spec 011 US5: 4-pattern v2 report
# ---------------------------------------------------------------------------

_PATTERN_LABEL: dict[str, str] = {
    "whole-same-week": "Whole Duplicate — Same Week",
    "scattered-same-week": "Scattered Duplicate — Same Week",
    "whole-different-week": "Whole Duplicate — Different Week",
    "scattered-different-week": "Scattered Duplicate — Different Week",
}

_PATTERN_ORDER = [
    "whole-same-week",
    "scattered-same-week",
    "whole-different-week",
    "scattered-different-week",
]

_TEMPLATES_DIR = Path(__file__).parent / "templates"


def _query_comparisons(db_path: Path, professor_id: str) -> list[dict[str, Any]]:
    """Fetch M-nC2 comparison rows for a professor.

    Args:
        db_path: SQLite content_reuse.db path.
        professor_id: Professor pool identifier.

    Returns:
        List of comparison result dicts.
    """
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM comparison_results "
            "WHERE matching_mode = 'M-nC2' AND professor_id = ? "
            "ORDER BY suspicion_score DESC",
            (professor_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _query_spans(db_path: Path, comparison_id: int) -> list[dict[str, Any]]:
    """Fetch match_spans for a comparison row.

    Args:
        db_path: SQLite content_reuse.db path.
        comparison_id: comparison_results.id.

    Returns:
        List of span dicts ordered by span_index.
    """
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM match_spans WHERE comparison_id = ? ORDER BY span_index",
            (comparison_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _query_professor(db_path: Path, professor_id: str) -> dict[str, Any] | None:
    """Fetch professor_pool row.

    Args:
        db_path: SQLite content_reuse.db path.
        professor_id: Professor pool identifier.

    Returns:
        Professor dict or None if not found.
    """
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT * FROM professor_pool WHERE professor_id = ?",
            (professor_id,),
        ).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def _query_phrase_whitelist(db_path: Path, professor_id: str) -> list[dict[str, Any]]:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM phrase_whitelist WHERE professor_id = ?", (professor_id,)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _query_pair_whitelist(db_path: Path) -> list[dict[str, Any]]:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM comparison_results WHERE review_status = 'FALSE_POSITIVE'"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _query_baseline(db_path: Path, professor_id: str) -> list[dict[str, Any]]:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM baseline_corpus WHERE professor_id = ?", (professor_id,)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _build_totals(comparisons: list[dict[str, Any]]) -> dict[str, int]:
    """Compute totals for the report header."""
    totals: dict[str, int] = {
        "total": len(comparisons),
        "whole_same": 0,
        "scattered_same": 0,
        "whole_diff": 0,
        "scattered_diff": 0,
        "layer_a_excluded": 0,
        "layer_c_demoted": 0,
        "layer_d_pair_excluded": 0,
        "layer_d_phrase_hits": 0,
        "layer_b_subtraction_events": 0,
    }
    pattern_map = {
        "whole-same-week": "whole_same",
        "scattered-same-week": "scattered_same",
        "whole-different-week": "whole_diff",
        "scattered-different-week": "scattered_diff",
    }
    for comp in comparisons:
        pat = comp.get("reuse_pattern") or ""
        if pat in pattern_map:
            totals[pattern_map[pat]] += 1

        attr_json = comp.get("layer_attribution") or "[]"
        try:
            attrs = json.loads(attr_json) if isinstance(attr_json, str) else attr_json
        except (json.JSONDecodeError, TypeError):
            attrs = []

        for attr in attrs:
            layer = attr.get("layer", "")
            action = attr.get("action", "")
            if layer == "A" and action == "excluded":
                totals["layer_a_excluded"] += 1
            elif layer == "C" and action == "demoted":
                totals["layer_c_demoted"] += 1
            elif layer == "D" and action == "subtracted":
                totals["layer_d_phrase_hits"] += 1
            elif layer == "B" and action == "subtracted":
                totals["layer_b_subtraction_events"] += 1

        if comp.get("review_status") == "FALSE_POSITIVE":
            totals["layer_d_pair_excluded"] += 1

    return totals


def _build_pair_data(
    comp: dict[str, Any],
    spans: list[dict[str, Any]],
) -> dict[str, Any]:
    """Build per-pair template data including time-axis chart."""
    from tube_scout.models.reuse_v2 import MatchSpan
    from tube_scout.visualization.time_axis_chart import render, render_to_base64_png

    match_spans = []
    for s in spans:
        try:
            match_spans.append(MatchSpan(
                start_a_seconds=float(s["start_a_seconds"]),
                end_a_seconds=float(s["end_a_seconds"]),
                start_b_seconds=float(s["start_b_seconds"]),
                end_b_seconds=float(s["end_b_seconds"]),
                length_seconds=float(s["length_seconds"]),
                matched_text_sample=s.get("matched_text_sample") or "",
                baseline_subtracted=bool(s.get("baseline_subtracted")),
                whitelisted=bool(s.get("whitelisted")),
            ))
        except Exception:
            pass

    duration_a = float(comp.get("i5_duration_diff_seconds") or 0) or 600.0
    duration_b = duration_a

    time_axis_chart_html = None
    time_axis_png_b64 = None

    pid = comp.get("comparison_id")
    if match_spans:
        try:
            fig = render(match_spans, duration_a=duration_a, duration_b=duration_b)
            time_axis_chart_html = fig.to_html(
                full_html=False,
                include_plotlyjs="cdn",
                config={"displayModeBar": False},
            )
        except Exception as exc:
            logger.warning("plotly render failed for pair %s: %s", pid, exc)

        try:
            time_axis_png_b64 = render_to_base64_png(
                match_spans, duration_a=duration_a, duration_b=duration_b
            )
        except Exception as exc:
            logger.warning("plotly png render failed for pair %s: %s", pid, exc)

    return {
        **comp,
        "time_axis_chart_html": time_axis_chart_html,
        "time_axis_png_b64": time_axis_png_b64,
        "matched_text_samples": [s.get("matched_text_sample") for s in spans if s.get("matched_text_sample")],
    }


def _generate_html(
    comparisons: list[dict[str, Any]],
    db_path: Path,
    professor: dict[str, Any],
    run_id: str,
    run_timestamp: str,
    output_path: Path,
) -> None:
    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATES_DIR)),
        autoescape=True,
    )
    tmpl = env.get_template("content_v2.html.j2")

    totals = _build_totals(comparisons)
    patterns: dict[str, list[dict[str, Any]]] = {pat: [] for pat in _PATTERN_ORDER}

    for comp in comparisons:
        pat = comp.get("reuse_pattern") or "whole-same-week"
        if pat not in patterns:
            patterns[pat] = []
        spans = _query_spans(db_path, comp["id"])
        pair_data = _build_pair_data(comp, spans)
        patterns[pat].append(pair_data)

    html = tmpl.render(
        professor=professor,
        run_id=run_id,
        run_timestamp=run_timestamp,
        totals=totals,
        patterns=patterns,
        pattern_order=_PATTERN_ORDER,
        pattern_label=_PATTERN_LABEL,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")


def _generate_xlsx(
    comparisons: list[dict[str, Any]],
    db_path: Path,
    professor_id: str,
    output_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)

    # Sheet 1: Summary
    ws_sum = wb.create_sheet("Summary")
    totals = _build_totals(comparisons)
    ws_sum.append(["Metric", "Value"])
    ws_sum["A1"].font = header_font
    ws_sum["B1"].font = header_font
    for k, v in totals.items():
        ws_sum.append([k, v])

    grade_counts = Counter(c.get("grade", "unknown") for c in comparisons)
    ws_sum.append([])
    ws_sum.append(["Grade Distribution", ""])
    for g, cnt in grade_counts.items():
        ws_sum.append([g, cnt])

    # Sheet 2: By Pattern
    ws_pat = wb.create_sheet("By Pattern")
    pat_headers = [
        "pattern", "source_video_id", "target_video_id",
        "suspicion_score", "grade", "review_status",
        "i2_cosine_similarity", "i6_longest_contiguous_seconds",
    ]
    ws_pat.append(pat_headers)
    for col, h in enumerate(pat_headers, 1):
        ws_pat.cell(row=1, column=col).font = header_font
    for comp in comparisons:
        ws_pat.append([_sanitize_cell(str(comp.get(h, ""))) for h in pat_headers])

    # Sheet 3: Whitelist
    ws_wl = wb.create_sheet("Whitelist")
    wl_headers = ["kind", "professor_id", "source_video_id", "target_video_id", "phrase_raw", "reason", "registered_by"]
    ws_wl.append(wl_headers)
    for col, h in enumerate(wl_headers, 1):
        ws_wl.cell(row=1, column=col).font = header_font
    for pair_wl in _query_pair_whitelist(db_path):
        ws_wl.append(["pair", "", pair_wl.get("source_video_id", ""), pair_wl.get("target_video_id", ""), "", "", ""])
    for phrase_wl in _query_phrase_whitelist(db_path, professor_id):
        ws_wl.append(["phrase", professor_id, "", "", phrase_wl.get("phrase_raw", ""), phrase_wl.get("reason", ""), phrase_wl.get("registered_by", "")])

    # Sheet 4: Baseline
    ws_base = wb.create_sheet("Baseline")
    base_headers = ["professor_id", "phrase_raw", "phrase_normalized", "occurrence_count", "registered_by"]
    ws_base.append(base_headers)
    for col, h in enumerate(base_headers, 1):
        ws_base.cell(row=1, column=col).font = header_font
    for row in _query_baseline(db_path, professor_id):
        ws_base.append([_sanitize_cell(str(row.get(h, ""))) for h in base_headers])

    # Sheet 5: Layer Attribution Audit
    ws_attr = wb.create_sheet("Layer Attribution")
    attr_headers = ["comparison_id", "source_video_id", "target_video_id", "layer", "action", "reason"]
    ws_attr.append(attr_headers)
    for col, h in enumerate(attr_headers, 1):
        ws_attr.cell(row=1, column=col).font = header_font
    for comp in comparisons:
        attr_json = comp.get("layer_attribution") or "[]"
        try:
            attrs = json.loads(attr_json) if isinstance(attr_json, str) else (attr_json or [])
        except (json.JSONDecodeError, TypeError):
            attrs = []
        if not attrs:
            ws_attr.append([comp.get("id"), comp.get("source_video_id"), comp.get("target_video_id"), "", "", ""])
        for attr in attrs:
            ws_attr.append([
                comp.get("id"),
                comp.get("source_video_id"),
                comp.get("target_video_id"),
                attr.get("layer", ""),
                attr.get("action", ""),
                attr.get("reason", ""),
            ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()


def _generate_json(
    comparisons: list[dict[str, Any]],
    db_path: Path,
    professor_id: str,
    run_id: str,
    run_timestamp: str,
    output_path: Path,
) -> None:
    pairs_with_spans = []
    for comp in comparisons:
        spans = _query_spans(db_path, comp["id"])
        pairs_with_spans.append({**comp, "match_spans": spans})

    report = {
        "run_id": run_id,
        "run_timestamp": run_timestamp,
        "professor_id": professor_id,
        "totals": _build_totals(comparisons),
        "comparisons": pairs_with_spans,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def generate_v2_report(
    project_dir: Path,
    professor_id: str,
    fmt: Literal["html", "xlsx", "json", "all"] = "html",
) -> dict[str, Path]:
    """Generate spec 011 v2 report(s) for one professor.

    Queries comparison_results filtered by matching_mode='M-nC2' AND
    professor_id, joins match_spans, computes totals (per pattern + Layer
    A/B/C/D counts).

    Output paths: 03_report/content/v2/{date}-{professor_id}-nc2.{ext}.

    Args:
        project_dir: Project root directory.
        professor_id: Professor pool identifier.
        fmt: Output format — 'html', 'xlsx', 'json', or 'all'.

    Returns:
        Dict mapping format extension to output Path.

    Raises:
        ValueError: If professor_id is not found in professor_pool or fmt is invalid.
        TypeError: If project_dir is not a Path.
    """
    if not isinstance(project_dir, Path):
        raise TypeError(f"project_dir must be a Path, got {type(project_dir).__name__}")
    if fmt not in ("html", "xlsx", "json", "all"):
        raise ValueError(f"fmt must be 'html', 'xlsx', 'json', or 'all', got {fmt!r}")

    db_path = project_dir / "02_analyze" / "content" / "content_reuse.db"
    professor = _query_professor(db_path, professor_id)
    if professor is None:
        raise ValueError(
            f"Professor '{professor_id}' not found in professor_pool. "
            "Register with 'tube-scout content professor map' first."
        )

    comparisons = _query_comparisons(db_path, professor_id)
    run_id = f"nc2-{professor_id}-{datetime.now(UTC).strftime('%Y%m%d-%H%M')}"
    run_timestamp = datetime.now(UTC).isoformat()
    date_str = datetime.now(UTC).strftime("%Y%m%d")

    out_dir = project_dir / "03_report" / "content" / "v2"
    stem = f"{date_str}-{professor_id}-nc2"

    result: dict[str, Path] = {}
    fmts = ["html", "xlsx", "json"] if fmt == "all" else [fmt]

    for f in fmts:
        path = out_dir / f"{stem}.{f}"
        if f == "html":
            _generate_html(comparisons, db_path, professor, run_id, run_timestamp, path)
        elif f == "xlsx":
            _generate_xlsx(comparisons, db_path, professor_id, path)
        elif f == "json":
            _generate_json(comparisons, db_path, professor_id, run_id, run_timestamp, path)
        result[f] = path

    return result
