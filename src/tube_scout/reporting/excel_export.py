"""Excel export for department reports (US4 — T045)."""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from tube_scout.models.report import (
    ComplianceMatrix,
    DepartmentOverview,
    ProfessorDetail,
)

# Conditional formatting fills
_GREEN_FILL = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
_RED_FILL = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
_HEADER_FONT = Font(bold=True)

_STATUS_FILLS = {
    "uploaded": _GREEN_FILL,
    "missing": _RED_FILL,
    "late": _YELLOW_FILL,
}


class ExcelExporter:
    """Export department report data to a multi-sheet Excel file.

    Sheets:
        1. 개요 — Department overview
        2. 교수별 상세 — Professor detail table
        3. 준수율 — Compliance matrix with conditional formatting
        4. 이상 탐지 — Validation findings (empty until US5)
    """

    def export(
        self,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
        output_path: Path,
        validation_findings: list[Any] | None = None,
    ) -> Path:
        """Export report data to an Excel file.

        Args:
            overview: Department overview metrics.
            professor_details: Per-professor detail list.
            compliance_entries: Compliance matrix entries.
            output_path: Path to write the xlsx file.
            validation_findings: Optional validation findings for sheet 4.

        Returns:
            Path to the generated Excel file.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: 개요
        self._write_overview_sheet(wb.active, overview)

        # Sheet 2: 교수별 상세
        ws2 = wb.create_sheet("교수별 상세")
        self._write_professor_detail_sheet(ws2, professor_details)

        # Sheet 3: 준수율
        ws3 = wb.create_sheet("준수율")
        self._write_compliance_sheet(ws3, compliance_entries)

        # Sheet 4: 이상 탐지
        ws4 = wb.create_sheet("이상 탐지")
        self._write_validation_sheet(ws4, validation_findings or [])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        wb.close()
        return output_path

    def _write_overview_sheet(
        self,
        ws: Any,
        overview: DepartmentOverview,
    ) -> None:
        """Write department overview to the first sheet.

        Args:
            ws: Worksheet object.
            overview: Department overview data.
        """
        ws.title = "개요"

        headers = ["항목", "값"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT

        rows = [
            ("채널 ID", overview.channel_id),
            ("채널명", overview.channel_name),
            ("연도", overview.year),
            ("학기", overview.semester),
            ("총 영상 수", overview.total_videos),
            ("총 교수 수", overview.total_professors),
            ("총 과목 수", overview.total_courses),
            ("총 시간 (시)", round(overview.total_duration_hours, 2)),
            ("총 조회수", overview.total_views),
            ("파싱 성공률", f"{overview.parse_success_rate:.1%}"),
        ]
        for row_idx, (label, value) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        # Auto-width
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _write_professor_detail_sheet(
        self,
        ws: Any,
        professor_details: list[ProfessorDetail],
    ) -> None:
        """Write professor detail table.

        Args:
            ws: Worksheet object.
            professor_details: List of professor details.
        """
        headers = [
            "교수명", "영상 수", "과목", "주차 커버리지",
            "차시 완결성", "평균 시간(분)", "총 조회수",
            "평균 조회수", "검증 오류 수",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT

        for row_idx, detail in enumerate(professor_details, 2):
            ws.cell(row=row_idx, column=1, value=detail.professor_name)
            ws.cell(row=row_idx, column=2, value=detail.video_count)
            ws.cell(row=row_idx, column=3, value=", ".join(detail.courses))
            ws.cell(row=row_idx, column=4, value=f"{detail.weekly_coverage:.1%}")
            ws.cell(row=row_idx, column=5, value=f"{detail.session_completeness:.1%}")
            ws.cell(row=row_idx, column=6, value=round(detail.avg_duration_minutes, 1))
            ws.cell(row=row_idx, column=7, value=detail.total_views)
            ws.cell(row=row_idx, column=8, value=round(detail.avg_views, 1))
            ws.cell(row=row_idx, column=9, value=detail.validation_error_count)

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_compliance_sheet(
        self,
        ws: Any,
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Write compliance matrix with conditional formatting.

        Args:
            ws: Worksheet object.
            compliance_entries: Compliance matrix entries.
        """
        # Header row: Professor + W1-W16 + Compliance Rate
        headers = ["교수명"] + [f"W{w}" for w in range(1, 17)] + ["준수율"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT

        for row_idx, entry in enumerate(compliance_entries, 2):
            ws.cell(row=row_idx, column=1, value=entry.professor_name)
            for week in range(1, 17):
                status = entry.week_statuses.get(week, "missing")
                cell = ws.cell(row=row_idx, column=week + 1, value=status)
                fill = _STATUS_FILLS.get(status)
                if fill:
                    cell.fill = fill
            ws.cell(
                row=row_idx, column=18,
                value=f"{entry.upload_deadline_compliance:.1%}",
            )

        # Auto-width
        ws.column_dimensions["A"].width = 15
        for col in range(2, 19):
            ws.column_dimensions[get_column_letter(col)].width = 10

    def _write_validation_sheet(
        self,
        ws: Any,
        validation_findings: list[Any],
    ) -> None:
        """Write validation findings (empty until US5 integration).

        Args:
            ws: Worksheet object.
            validation_findings: List of ValidationFinding objects.
        """
        headers = ["규칙 ID", "심각도", "영상 ID", "교수명", "설명", "상세"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT

        for row_idx, finding in enumerate(validation_findings, 2):
            ws.cell(row=row_idx, column=1, value=finding.rule_id)
            ws.cell(row=row_idx, column=2, value=finding.severity)
            ws.cell(row=row_idx, column=3, value=", ".join(finding.video_ids))
            ws.cell(row=row_idx, column=4, value=finding.professor or "")
            ws.cell(row=row_idx, column=5, value=finding.description)
            ws.cell(row=row_idx, column=6, value=str(finding.details))

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
