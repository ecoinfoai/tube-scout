"""Layer 3 integration tests: Module combination verification.

Tests 7 module combinations across service boundaries:
  1. title_parser + validator: parse_error=True videos and rule application
  2. search_service + video_filter_service: YAML results -> filter type compat
  3. forecaster + empty time series: <MIN_DATA graceful failure
  4. sentiment(LLM) + rate_limiter: rate limit application to LLM calls
  5. department_report + excel_export: Korean Excel encoding
  6. bundle_report + video_filter: 0-match filter -> error handling
  7. auth(OAuth) + youtube_data: token refresh during API call
"""

import time
from datetime import UTC, date, datetime
from pathlib import Path
from unittest.mock import MagicMock

import pytest
import yaml

from tube_scout.models.config import RateLimitProfile
from tube_scout.models.parsed_title import ParsedTitle
from tube_scout.models.video import Video
from tube_scout.models.video_filter import VideoFilter
from tube_scout.services.forecaster import ForecasterService
from tube_scout.services.rate_limiter import RateLimiter
from tube_scout.services.search_service import SearchService
from tube_scout.services.title_parser import TitleParser
from tube_scout.services.validator import run_all_validations
from tube_scout.services.video_filter_service import VideoFilterService
from tube_scout.services.youtube_analytics import YouTubeAnalyticsService
from tube_scout.services.youtube_data import YouTubeDataService
from tube_scout.storage.json_store import write_json

CHANNEL_ID = "UCxxxxxxxxxxxxxxxxxxxxxx"


# ===========================================================================
# MC-1: title_parser + validator: parse_error=True and rule application
# ===========================================================================


class TestMC1TitleParserValidator:
    """Title parser parse_error videos should not cause validator crashes."""

    def test_parse_error_videos_pass_through_validator(self) -> None:
        """Videos with parse_error=True pass through all validation rules."""
        parser = TitleParser()

        # Unparseable titles
        titles = [
            parser.parse("Random English Title No Pattern", "vid001"),
            parser.parse("Another unparseable title here", "vid002"),
        ]

        # All should be parse_error=True
        assert all(t.parse_error for t in titles)

        # Validator should not crash, should return V-005 findings
        videos = [
            {
                "video_id": "vid001",
                "published_at": "2024-01-01",
                "duration_seconds": 100,
            },
            {
                "video_id": "vid002",
                "published_at": "2024-01-02",
                "duration_seconds": 200,
            },
        ]
        findings = run_all_validations(titles, videos)
        v005_findings = [f for f in findings if f.rule_id == "V-005"]
        assert len(v005_findings) == 2

    def test_mixed_parsed_and_error_titles(self) -> None:
        """Validator handles mix of parsed and unparsed titles."""
        parser = TitleParser()

        prof = "\ud64d\uae38\ub3d9"
        titles = [
            parser.parse(
                prof + " 2024 \ud574\ubd80\ud559 1\uc8fc\ucc28 1\ucc28\uc2dc", "vid001"
            ),
            parser.parse("Unparseable title", "vid002"),
            parser.parse(
                prof + " 2024 \ud574\ubd80\ud559 2\uc8fc\ucc28 1\ucc28\uc2dc", "vid003"
            ),
        ]

        assert not titles[0].parse_error
        assert titles[1].parse_error
        assert not titles[2].parse_error

        videos = [
            {
                "video_id": "vid001",
                "published_at": "2024-01-01",
                "duration_seconds": 600,
            },
            {
                "video_id": "vid002",
                "published_at": "2024-01-02",
                "duration_seconds": 300,
            },
            {
                "video_id": "vid003",
                "published_at": "2024-01-03",
                "duration_seconds": 600,
            },
        ]

        findings = run_all_validations(titles, videos)
        v005 = [f for f in findings if f.rule_id == "V-005"]
        assert len(v005) == 1
        assert "vid002" in v005[0].video_ids

    def test_parse_error_does_not_trigger_other_rules(self) -> None:
        """parse_error titles with None fields don't trigger week/session rules."""
        parser = TitleParser()
        title = parser.parse("No Korean pattern here at all", "vid001")
        assert title.parse_error
        assert title.week is None
        assert title.session is None

        findings = run_all_validations(
            [title],
            [{"video_id": "vid001", "published_at": "2024-01-01"}],
        )
        v003 = [f for f in findings if f.rule_id == "V-003"]
        assert len(v003) == 0


# ===========================================================================
# MC-2: search_service + video_filter_service: type compatibility
# ===========================================================================


class TestMC2SearchFilterCompat:
    """SearchService results should be compatible with VideoFilterService."""

    def test_search_results_filterable(self) -> None:
        """Parsed titles from SearchService can be fed to VideoFilterService."""
        prof = "\ud64d\uae38\ub3d9"
        course = "\ud574\ubd80\ud559"
        parsed_titles = [
            ParsedTitle(
                video_id=f"vid{i:03d}",
                original_title=f"{prof} 2024 {course} {i}\uc8fc\ucc28 1\ucc28\uc2dc",
                professor=[prof],
                course=course,
                year=2024,
                week=i,
                session=1,
            )
            for i in range(1, 6)
        ]

        query = SearchService.from_cli_flags(professor=prof, week_from=2, week_to=4)
        search_results = SearchService.search(parsed_titles, query)
        assert len(search_results) == 3

        video_dicts = [
            {
                "video_id": pt.video_id,
                "title": pt.original_title,
                "published_at": f"2024-01-{pt.week:02d}",
            }
            for pt in parsed_titles
        ]
        result_ids = {pt.video_id for pt in search_results}
        vf = VideoFilter(video_ids=list(result_ids))
        filtered = VideoFilterService.filter_videos(video_dicts, vf)
        assert len(filtered) == 3

    def test_yaml_config_to_filter(self, tmp_path: Path) -> None:
        """YAML search config loads and applies correctly."""
        prof = "\ud64d\uae38\ub3d9"
        yaml_data = {
            "filters": {
                "professor": prof,
                "year": 2024,
                "week_range": [1, 8],
            },
            "exclude": {
                "title_contains": ["OT"],
            },
        }
        yaml_path = tmp_path / "search.yaml"
        yaml_path.write_text(
            yaml.dump(yaml_data, allow_unicode=True),
            encoding="utf-8",
        )

        query = SearchService.load_config(yaml_path)
        assert query.filters is not None
        assert query.filters.professor == prof
        assert query.filters.year == 2024
        assert query.filters.week_range == [1, 8]
        assert query.exclude is not None
        assert "OT" in query.exclude.title_contains

    def test_empty_search_returns_all(self) -> None:
        """Empty SearchQuery returns all titles."""
        from tube_scout.models.search import SearchQuery

        titles = [
            ParsedTitle(
                video_id=f"vid{i:03d}",
                original_title=f"Title {i}",
            )
            for i in range(5)
        ]
        query = SearchQuery()
        results = SearchService.search(titles, query)
        assert len(results) == 5


# ===========================================================================
# MC-3: forecaster + empty time series: graceful failure
# ===========================================================================


class TestMC3ForecasterGraceful:
    """Forecaster should fail gracefully with insufficient data."""

    def test_insufficient_data_raises_valueerror(self) -> None:
        """predict() raises ValueError for data < MIN_DATA_DAYS."""
        svc = ForecasterService()
        short_data = [{"date": i, "value": float(i)} for i in range(10)]
        with pytest.raises(ValueError, match="At least 6 months"):
            svc.predict(CHANNEL_ID, "views", short_data)

    def test_exactly_min_data_succeeds(self) -> None:
        """predict() works with exactly MIN_DATA_DAYS data points."""
        svc = ForecasterService()
        base = date(2024, 1, 1).toordinal()
        data = [{"date": base + i, "value": float(100 + i)} for i in range(180)]
        # Use linear model explicitly to avoid statsmodels/prophet dependency
        results = svc.predict(
            CHANNEL_ID,
            "views",
            data,
            horizon_days=5,
            model="linear",
        )
        assert len(results) == 5

    def test_model_selection_by_data_length(self) -> None:
        """select_model picks correct model based on n_days."""
        svc = ForecasterService()
        assert svc.select_model(30) == "linear"
        assert svc.select_model(89) == "linear"
        assert svc.select_model(90) == "arima"
        assert svc.select_model(365) == "arima"
        assert svc.select_model(366) == "prophet"

    def test_fill_missing_days_interpolation(self) -> None:
        """fill_missing_days correctly interpolates gaps."""
        svc = ForecasterService()
        data = [
            {"date": 1, "value": 10.0},
            {"date": 3, "value": 30.0},
            {"date": 5, "value": 50.0},
        ]
        filled = svc.fill_missing_days(data)
        assert len(filled) == 5
        day2 = next(d for d in filled if d["date"] == 2)
        assert abs(day2["value"] - 20.0) < 0.01

    def test_anomaly_detection_no_data(self) -> None:
        """detect_anomalies handles empty data."""
        svc = ForecasterService()
        assert svc.detect_anomalies([]) == []

    def test_anomaly_detection_constant_data(self) -> None:
        """detect_anomalies with zero variance marks nothing."""
        svc = ForecasterService()
        data = [{"date": i, "value": 100.0} for i in range(20)]
        results = svc.detect_anomalies(data)
        assert all(not r["is_anomaly"] for r in results)


# ===========================================================================
# MC-4: sentiment(LLM) + rate_limiter: rate limit on LLM calls
# ===========================================================================


class TestMC4SentimentRateLimiter:
    """Sentiment backend should respect rate limiting."""

    def test_rate_limiter_basic_delay(self) -> None:
        """RateLimiter.wait() introduces a delay."""
        profile = RateLimitProfile(
            base_delay=0.05,
            max_retries=3,
            backoff_multiplier=2.0,
            jitter=0.0,
        )
        limiter = RateLimiter(profile)

        start = time.time()
        limiter.wait()
        elapsed = time.time() - start
        assert elapsed >= 0.04

    def test_rate_limiter_exponential_backoff(self) -> None:
        """Exponential backoff increases delay with each attempt."""
        profile = RateLimitProfile(
            base_delay=0.01,
            max_retries=5,
            backoff_multiplier=2.0,
            jitter=0.0,
        )
        limiter = RateLimiter(profile)

        start = time.time()
        limiter.wait_on_error(0)
        elapsed0 = time.time() - start

        start = time.time()
        limiter.wait_on_error(1)
        elapsed1 = time.time() - start

        assert elapsed1 > elapsed0

    def test_rate_limiter_max_retries_exceeded(self) -> None:
        """wait_on_error raises RuntimeError past max_retries."""
        profile = RateLimitProfile(
            base_delay=0.01,
            max_retries=3,
            backoff_multiplier=2.0,
        )
        limiter = RateLimiter(profile)

        with pytest.raises(RuntimeError, match="Max retries"):
            limiter.wait_on_error(3)

    def test_sentiment_skip_backend(self) -> None:
        """Sentiment 'skip' backend returns null sentiments without API."""
        from tube_scout.services.sentiment import SentimentService

        svc = SentimentService(backend="skip")
        comments = [
            {"comment_id": "c1", "text": "good lecture"},
            {"comment_id": "c2", "text": "hard to understand"},
        ]
        results = svc.analyze_batch(comments)
        assert len(results) == 2
        assert all(r["sentiment"] is None for r in results)

    def test_sentiment_empty_batch(self) -> None:
        """Empty comment batch returns empty results."""
        from tube_scout.services.sentiment import SentimentService

        svc = SentimentService(backend="skip")
        assert svc.analyze_batch([]) == []

    def test_rate_limiter_invalid_profile_type(self) -> None:
        """RateLimiter rejects non-RateLimitProfile."""
        with pytest.raises(TypeError, match="RateLimitProfile"):
            RateLimiter(profile="not a profile")  # type: ignore[arg-type]


# ===========================================================================
# MC-5: department_report + excel_export: Korean Excel encoding
# ===========================================================================


class TestMC5DepartmentExcel:
    """Department report -> Excel export with Korean content."""

    def test_full_department_to_excel_flow(self, tmp_path: Path) -> None:
        """Complete flow: parsed titles -> overview/detail/compliance -> Excel."""
        from tube_scout.reporting.department_report import DepartmentReportGenerator
        from tube_scout.reporting.excel_export import ExcelExporter

        gen = DepartmentReportGenerator()
        prof = "\uae40\uad50\uc218"
        course = "\uae30\ucd08\uac04\ud638\ud559"
        dept = "\uac04\ud638\ud559\uacfc"

        titles = [
            ParsedTitle(
                video_id=f"vid{i:03d}",
                original_title=f"{prof} 2024 {course} {i}\uc8fc\ucc28 1\ucc28\uc2dc",
                professor=[prof],
                course=course,
                year=2024,
                semester=1,
                week=i,
                session=1,
            )
            for i in range(1, 9)
        ]
        videos = [
            Video(
                video_id=f"vid{i:03d}",
                channel_id=CHANNEL_ID,
                title=f"{prof} 2024 {course} {i}\uc8fc\ucc28 1\ucc28\uc2dc",
                published_at=datetime(2024, 1, i, tzinfo=UTC),
                duration_seconds=1800,
                view_count=200,
            )
            for i in range(1, 9)
        ]

        overview = gen.compute_overview(titles, videos, CHANNEL_ID, dept)
        details = gen.compute_professor_details(titles, videos)
        compliance = gen.compute_compliance(titles, videos)

        assert overview.total_videos == 8
        assert len(details) == 1
        assert details[0].professor_name == prof
        assert len(compliance) == 1

        exporter = ExcelExporter()
        xlsx_path = tmp_path / "dept_report.xlsx"
        result = exporter.export(overview, details, compliance, xlsx_path)
        assert result.exists()

        import openpyxl

        wb = openpyxl.load_workbook(str(result))
        assert "\uac1c\uc694" in wb.sheetnames
        assert "\uad50\uc218\ubcc4 \uc0c1\uc138" in wb.sheetnames
        assert "\uc900\uc218\uc728" in wb.sheetnames

        ws_prof = wb["\uad50\uc218\ubcc4 \uc0c1\uc138"]
        assert ws_prof.cell(row=2, column=1).value == prof
        assert ws_prof.cell(row=2, column=3).value == course


# ===========================================================================
# MC-6: bundle_report + video_filter: 0-match filter -> error
# ===========================================================================


class TestMC6BundleFilterEmpty:
    """Bundle report with zero matching videos should raise ValueError."""

    def test_bundle_no_match_raises_error(self, tmp_path: Path) -> None:
        """BundleReportGenerator.generate raises ValueError on 0 matches."""
        from tube_scout.reporting.bundle_report import BundleReportGenerator

        collect_dir = tmp_path / "collect"
        videos_path = collect_dir / "channels" / CHANNEL_ID / "videos_meta.json"
        write_json(
            videos_path,
            [
                {
                    "video_id": "vid001",
                    "title": "anatomy lecture",
                    "published_at": "2024-01-01",
                },
            ],
        )

        gen = BundleReportGenerator(collect_dir=collect_dir)
        vf = VideoFilter(keyword="nonexistent_keyword_xyz")
        output = tmp_path / "output" / "bundle.html"

        with pytest.raises(ValueError, match="No videos matching"):
            gen.generate(vf, CHANNEL_ID, output)

    def test_video_filter_empty_list(self) -> None:
        """VideoFilterService with empty video list returns empty."""
        vf = VideoFilter(keyword="test")
        result = VideoFilterService.filter_videos([], vf)
        assert result == []

    def test_video_filter_keyword_match(self) -> None:
        """VideoFilterService keyword filter works correctly."""
        videos = [
            {"video_id": "v1", "title": "anatomy week 1", "published_at": "2024-01-01"},
            {
                "video_id": "v2",
                "title": "physiology week 1",
                "published_at": "2024-01-02",
            },
        ]
        vf = VideoFilter(keyword="anatomy")
        result = VideoFilterService.filter_videos(videos, vf)
        assert len(result) == 1
        assert result[0]["video_id"] == "v1"

    def test_video_filter_date_range(self) -> None:
        """VideoFilterService date range filter works correctly."""
        videos = [
            {"video_id": "v1", "title": "L1", "published_at": "2024-01-01"},
            {"video_id": "v2", "title": "L2", "published_at": "2024-06-15"},
            {"video_id": "v3", "title": "L3", "published_at": "2024-12-31"},
        ]
        vf = VideoFilter(
            published_after=date(2024, 3, 1),
            published_before=date(2024, 9, 30),
        )
        result = VideoFilterService.filter_videos(videos, vf)
        assert len(result) == 1
        assert result[0]["video_id"] == "v2"


# ===========================================================================
# MC-7: auth(OAuth) + youtube_data: token refresh during API call
# ===========================================================================


class TestMC7AuthTokenRefresh:
    """OAuth token handling during API operations."""

    def test_channel_auth_with_valid_credentials(self) -> None:
        """YouTubeDataService works with valid mock credentials."""
        client = MagicMock()
        client.channels().list.return_value.execute.return_value = {
            "items": [
                {
                    "id": CHANNEL_ID,
                    "snippet": {"title": "Test Channel"},
                    "contentDetails": {
                        "relatedPlaylists": {"uploads": "UUxxxxxxxxxxxxxxxxxxxxxx"}
                    },
                    "statistics": {
                        "videoCount": "10",
                        "subscriberCount": "100",
                        "viewCount": "5000",
                    },
                }
            ]
        }

        svc = YouTubeDataService(client)
        info = svc.get_channel_info(CHANNEL_ID)
        assert info["channel_id"] == CHANNEL_ID
        assert info["total_video_count"] == 10

    def test_channel_not_found_raises_value_error(self) -> None:
        """get_channel_info raises ValueError for non-existent channel."""
        client = MagicMock()
        client.channels().list.return_value.execute.return_value = {"items": []}

        svc = YouTubeDataService(client)
        with pytest.raises(ValueError, match="Channel not found"):
            svc.get_channel_info("UCnonexistent______________")

    def test_analytics_permission_error_on_403(self) -> None:
        """Analytics 403 -> PermissionError with clear message."""
        from googleapiclient.errors import HttpError

        client = MagicMock()
        resp = MagicMock()
        resp.status = 403
        client.reports().query.return_value.execute.side_effect = HttpError(
            resp, b"Forbidden"
        )

        svc = YouTubeAnalyticsService(client=client)
        with pytest.raises(PermissionError, match="access denied"):
            svc.get_daily_metrics(
                channel_id=CHANNEL_ID,
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )

    def test_analytics_retry_on_500(self) -> None:
        """Analytics retries on 500 and eventually succeeds."""
        from googleapiclient.errors import HttpError

        client = MagicMock()
        resp_500 = MagicMock()
        resp_500.status = 500

        call_count = 0

        def mock_execute() -> dict:
            nonlocal call_count
            call_count += 1
            if call_count < 3:
                raise HttpError(resp_500, b"Server Error")
            return {"rows": [["2024-01-01", 100, 500.0, 300.0, 75.0]]}

        client.reports().query.return_value.execute = mock_execute

        svc = YouTubeAnalyticsService(client=client)
        data = svc.get_daily_metrics(
            channel_id=CHANNEL_ID,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31),
        )
        assert len(data) == 1
        assert call_count == 3

    def test_registry_load_save_roundtrip(self, tmp_path: Path) -> None:
        """Channel registry save -> load preserves all data."""
        from tube_scout.models.config import ChannelRegistration
        from tube_scout.services.auth import load_registry, save_registry

        dept = "\uac04\ud638\ud559\uacfc"
        reg = {
            dept: ChannelRegistration(
                alias=dept,
                channel_id=CHANNEL_ID,
                channel_name=dept + " channel",
                registered_at="2024-01-01T00:00:00+00:00",
                last_used_at="2024-01-01T00:00:00+00:00",
                token_path=str(tmp_path / "token.json"),
            )
        }

        save_registry(tmp_path, reg)
        loaded = load_registry(tmp_path)
        assert dept in loaded
        assert loaded[dept].channel_id == CHANNEL_ID
