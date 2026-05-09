"""Integration tests for v2 report generation e2e (T065 RED).

SC-006: 4 패턴 섞인 합성 데이터 → HTML 4 섹션 + Excel 5 시트 + JSON dump 정상.
Tests CLI 'report content --professor --format' command wiring.
"""

import json
import sqlite3
from pathlib import Path

import pytest
from typer.testing import CliRunner

from tests.fixtures.spec011.fixture_db import build_clean_v2_db


def _make_full_project(tmp_path: Path) -> tuple[Path, Path]:
    """Create project with professor + 4-pattern comparison rows + match_spans."""
    db_dir = tmp_path / "02_analyze" / "content"
    db_dir.mkdir(parents=True)
    db_path = db_dir / "content_reuse.db"
    build_clean_v2_db(db_path)

    conn = sqlite3.connect(str(db_path))
    conn.execute(
        "INSERT OR IGNORE INTO professor_pool (professor_id, display_name, created_at, created_by) "
        "VALUES ('prof-e2e-rpt', 'E2E Report Prof', '2026-01-01T00:00:00', 'system')"
    )
    patterns = [
        "whole-same-week",
        "scattered-same-week",
        "whole-different-week",
        "scattered-different-week",
    ]
    comp_ids = []
    for i, pat in enumerate(patterns):
        cur = conn.execute(
            "INSERT OR IGNORE INTO comparison_results "
            "(source_video_id, target_video_id, matching_mode, professor_id, "
            "review_status, reuse_pattern, suspicion_score, grade, "
            "i6_longest_contiguous_seconds, created_at) "
            "VALUES (?, ?, 'M-nC2', 'prof-e2e-rpt', 'UNREVIEWED', ?, ?, 'high', 300.0, '2026-01-01T00:00:00')",
            (f"e2e-src-{i}", f"e2e-tgt-{i}", pat, 80.0 - i * 5),
        )
        comp_ids.append(cur.lastrowid)

    # Insert match_spans for each comparison
    for comp_id in comp_ids:
        conn.execute(
            "INSERT OR IGNORE INTO match_spans "
            "(comparison_id, span_index, start_a_seconds, end_a_seconds, "
            "start_b_seconds, end_b_seconds, length_seconds, matched_text_sample) "
            "VALUES (?, 0, 0.0, 300.0, 0.0, 300.0, 300.0, '강의 본론 내용')",
            (comp_id,),
        )

    conn.commit()
    conn.close()
    return tmp_path, db_path


def test_sc006_html_four_pattern_sections(tmp_path: Path) -> None:
    """SC-006a: HTML report contains 4 pattern sections."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_full_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-e2e-rpt", fmt="html")

    content = paths["html"].read_text(encoding="utf-8")
    for pat in ["whole-same-week", "scattered-same-week", "whole-different-week", "scattered-different-week"]:
        assert pat in content, f"Pattern section '{pat}' not in HTML"


def test_sc006_xlsx_five_sheets(tmp_path: Path) -> None:
    """SC-006b: Excel report has 5 required sheets."""
    from tube_scout.reporting.content_report import generate_v2_report

    import openpyxl

    project_dir, _ = _make_full_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-e2e-rpt", fmt="xlsx")

    wb = openpyxl.load_workbook(str(paths["xlsx"]))
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    required = ["summary", "by pattern", "whitelist", "baseline", "layer attribution"]
    for req in required:
        assert any(req in s for s in sheet_names_lower), (
            f"Required sheet '{req}' not found in {wb.sheetnames}"
        )


def test_sc006_json_has_comparisons(tmp_path: Path) -> None:
    """SC-006c: JSON report contains comparison data."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_full_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-e2e-rpt", fmt="json")

    data = json.loads(paths["json"].read_text(encoding="utf-8"))
    if isinstance(data, dict):
        comparisons = data.get("comparisons") or data.get("pairs") or []
    else:
        comparisons = data
    assert len(comparisons) >= 4, f"Expected 4 comparisons in JSON, got {len(comparisons)}"


def test_sc006_all_format_creates_three_files(tmp_path: Path) -> None:
    """SC-006d: fmt='all' creates html + xlsx + json."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_full_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-e2e-rpt", fmt="all")

    assert set(paths.keys()) >= {"html", "xlsx", "json"}
    for key, path in paths.items():
        assert path.exists(), f"'{key}' file not created: {path}"


def test_sc006_cli_report_content_command(tmp_path: Path) -> None:
    """CLI 'report content --professor --format html' exits 0 and creates file."""
    from tube_scout.cli.main import app
    import tube_scout.cli.content as _content_mod

    project_dir, _ = _make_full_project(tmp_path)

    original_flag = _content_mod._SPEC011_MIGRATED
    _content_mod._SPEC011_MIGRATED = True
    try:
        runner = CliRunner()
        result = runner.invoke(
            app,
            [
                "report", "content",
                "--project", str(project_dir),
                "--professor", "prof-e2e-rpt",
                "--format", "html",
            ],
            catch_exceptions=True,
        )
    finally:
        _content_mod._SPEC011_MIGRATED = original_flag

    assert result.exit_code == 0, (
        f"Expected exit 0, got {result.exit_code}. "
        f"Output: {result.output!r}. Exception: {result.exception!r}"
    )
    report_dir = project_dir / "03_report" / "content" / "v2"
    html_files = list(report_dir.glob("*.html")) if report_dir.exists() else []
    assert html_files, f"No HTML file created under {report_dir}"


def test_sc006_cli_missing_professor_exits_nonzero(tmp_path: Path) -> None:
    """CLI 'report content' without --professor exits non-zero."""
    from tube_scout.cli.main import app

    project_dir, _ = _make_full_project(tmp_path)

    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "report", "content",
            "--project", str(project_dir),
            "--format", "html",
        ],
        catch_exceptions=True,
    )
    assert result.exit_code != 0
