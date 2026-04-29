"""Tests for content quality report generator."""

import json
from pathlib import Path

import pytest

from tube_scout.reporting.content_report import ContentReportGenerator


@pytest.fixture()
def sample_comparisons() -> list[dict]:
    """Sample comparison results for testing."""
    return [
        {
            "id": 1, "source_video_id": "v1", "target_video_id": "v2",
            "professor": "Kim", "course": "Math", "week": 1, "session": 1,
            "year_from": 2025, "year_to": 2026,
            "i1_hash_match": 1, "i2_cosine_similarity": 0.95,
            "i3_change_rate": 0.05, "i4_new_term_count": 2,
            "i5_duration_diff_seconds": 5.0,
            "suspicion_score": 90.0, "grade": "critical",
            "review_status": "UNREVIEWED",
        },
        {
            "id": 2, "source_video_id": "v3", "target_video_id": "v4",
            "professor": "Lee", "course": "Physics", "week": 1, "session": 1,
            "year_from": 2025, "year_to": 2026,
            "i1_hash_match": 0, "i2_cosine_similarity": 0.60,
            "i3_change_rate": 0.40, "i4_new_term_count": 15,
            "i5_duration_diff_seconds": 120.0,
            "suspicion_score": 35.0, "grade": "normal",
            "review_status": "UNREVIEWED",
        },
    ]


@pytest.fixture()
def sample_quality() -> list[dict]:
    """Sample quality results for testing."""
    return [
        {
            "video_id": "v1", "q001_voice_present": 1, "q002_min_duration": 1,
            "q003_course_relevance": 0.3, "q004_silence_ratio": 0.15,
            "q005_speech_density": 350.0, "pass_count": 5,
        },
        {
            "video_id": "v2", "q001_voice_present": 1, "q002_min_duration": 0,
            "q003_course_relevance": None, "q004_silence_ratio": 0.5,
            "q005_speech_density": 100.0, "pass_count": 1,
        },
    ]


class TestContentReportJSON:
    """Tests for JSON report generation."""

    def test_generate_json(
        self, tmp_path: Path, sample_comparisons: list, sample_quality: list
    ) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.json"
        gen.generate_json(sample_comparisons, sample_quality, output)
        assert output.exists()
        data = json.loads(output.read_text())
        assert "summary" in data
        assert "comparisons" in data
        assert "quality_results" in data

    def test_json_summary_grade_counts(
        self, tmp_path: Path, sample_comparisons: list, sample_quality: list
    ) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.json"
        gen.generate_json(sample_comparisons, sample_quality, output)
        data = json.loads(output.read_text())
        summary = data["summary"]
        assert summary["total_comparisons"] == 2
        assert summary["grade_counts"]["critical"] == 1
        assert summary["grade_counts"]["normal"] == 1

    def test_json_empty_data(self, tmp_path: Path) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.json"
        gen.generate_json([], [], output)
        assert output.exists()
        data = json.loads(output.read_text())
        assert data["summary"]["total_comparisons"] == 0


class TestContentReportHTML:
    """Tests for HTML report generation."""

    def test_generate_html(
        self, tmp_path: Path, sample_comparisons: list, sample_quality: list
    ) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.html"
        gen.generate_html(sample_comparisons, sample_quality, output)
        assert output.exists()
        content = output.read_text()
        assert "<html" in content
        assert "critical" in content.lower()

    def test_html_contains_professor(
        self, tmp_path: Path, sample_comparisons: list, sample_quality: list
    ) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.html"
        gen.generate_html(sample_comparisons, sample_quality, output)
        content = output.read_text()
        assert "Kim" in content

    def test_html_empty_data(self, tmp_path: Path) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.html"
        gen.generate_html([], [], output)
        assert output.exists()


class TestContentReportExcel:
    """Tests for Excel report generation."""

    def test_generate_xlsx(
        self, tmp_path: Path, sample_comparisons: list, sample_quality: list
    ) -> None:
        gen = ContentReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate_xlsx(sample_comparisons, sample_quality, output)
        assert output.exists()
        assert output.stat().st_size > 0

    def test_xlsx_has_sheets(
        self, tmp_path: Path, sample_comparisons: list, sample_quality: list
    ) -> None:
        import openpyxl

        gen = ContentReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate_xlsx(sample_comparisons, sample_quality, output)
        wb = openpyxl.load_workbook(output)
        sheet_names = wb.sheetnames
        assert "Suspicion Summary" in sheet_names
        assert "Quality Results" in sheet_names
        wb.close()

    def test_xlsx_sanitizes_formulas(
        self, tmp_path: Path, sample_quality: list
    ) -> None:
        """Excel injection should be prevented."""
        comparisons = [
            {
                "id": 1, "source_video_id": "v1", "target_video_id": "v2",
                "professor": "=CMD()", "course": "+DANGEROUS",
                "week": 1, "session": 1,
                "year_from": 2025, "year_to": 2026,
                "i1_hash_match": 0, "i2_cosine_similarity": 0.5,
                "i3_change_rate": 0.5, "i4_new_term_count": 5,
                "i5_duration_diff_seconds": 30.0,
                "suspicion_score": 50.0, "grade": "moderate",
                "review_status": "UNREVIEWED",
            },
        ]
        gen = ContentReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate_xlsx(comparisons, sample_quality, output)

        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb["Suspicion Summary"]
        # Check that formula-like values are sanitized
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell:
                    assert cell[0] != "=" or cell.startswith("'=")
                    assert cell[0] != "+" or cell.startswith("'+")
        wb.close()
