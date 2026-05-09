"""Unit tests for generate_v2_report (T063 RED).

Tests content_report.generate_v2_report: 4-pattern HTML, Excel 5-sheet,
JSON dump, professor_id validation, and fmt validation.
"""

import json
import sqlite3
from pathlib import Path

import pytest

from tests.fixtures.spec011.fixture_db import build_clean_v2_db


def _make_project(tmp_path: Path) -> tuple[Path, Path]:
    """Return (project_dir, db_path) with professor + 4-pattern comparison rows."""
    db_dir = tmp_path / "02_analyze" / "content"
    db_dir.mkdir(parents=True)
    db_path = db_dir / "content_reuse.db"
    build_clean_v2_db(db_path)

    conn = sqlite3.connect(str(db_path))
    conn.execute(
        "INSERT OR IGNORE INTO professor_pool (professor_id, display_name, created_at, created_by) "
        "VALUES ('prof-report', 'Report Prof', '2026-01-01T00:00:00', 'system')"
    )
    # Insert 4 comparison rows with different reuse_patterns
    patterns = [
        "whole-same-week",
        "scattered-same-week",
        "whole-different-week",
        "scattered-different-week",
    ]
    for i, pat in enumerate(patterns):
        conn.execute(
            "INSERT OR IGNORE INTO comparison_results "
            "(source_video_id, target_video_id, matching_mode, professor_id, "
            "review_status, reuse_pattern, suspicion_score, grade, created_at) "
            "VALUES (?, ?, 'M-nC2', 'prof-report', 'UNREVIEWED', ?, ?, 'high', '2026-01-01T00:00:00')",
            (f"src-{i}", f"tgt-{i}", pat, 80.0 - i * 5),
        )
    conn.commit()
    conn.close()
    return tmp_path, db_path


def test_generate_v2_report_html_creates_file(tmp_path: Path) -> None:
    """generate_v2_report with fmt='html' creates an HTML file under 03_report/content/v2/."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-report", fmt="html")

    assert "html" in paths
    html_path = paths["html"]
    assert html_path.exists()
    assert html_path.suffix == ".html"
    content = html_path.read_text(encoding="utf-8")
    # 4 patterns must appear
    assert "whole-same-week" in content or "whole" in content.lower()


def test_generate_v2_report_html_has_four_pattern_sections(tmp_path: Path) -> None:
    """HTML report contains sections for all 4 reuse patterns."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-report", fmt="html")

    content = paths["html"].read_text(encoding="utf-8")
    for pat in ["whole-same-week", "scattered-same-week", "whole-different-week", "scattered-different-week"]:
        assert pat in content, f"Pattern section '{pat}' missing from HTML"


def test_generate_v2_report_xlsx_creates_five_sheets(tmp_path: Path) -> None:
    """generate_v2_report with fmt='xlsx' creates Excel with required sheets."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-report", fmt="xlsx")

    assert "xlsx" in paths
    xlsx_path = paths["xlsx"]
    assert xlsx_path.exists()

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path))
    sheet_names = [s.lower() for s in wb.sheetnames]
    for expected in ["summary", "by pattern", "whitelist", "baseline", "layer attribution"]:
        assert any(expected in s for s in sheet_names), (
            f"Sheet '{expected}' missing. Sheets: {wb.sheetnames}"
        )


def test_generate_v2_report_json_creates_valid_json(tmp_path: Path) -> None:
    """generate_v2_report with fmt='json' creates parseable JSON with comparisons."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-report", fmt="json")

    assert "json" in paths
    json_path = paths["json"]
    assert json_path.exists()
    data = json.loads(json_path.read_text(encoding="utf-8"))
    assert "comparisons" in data or "pairs" in data or isinstance(data, list)


def test_generate_v2_report_all_creates_three_files(tmp_path: Path) -> None:
    """generate_v2_report with fmt='all' creates html, xlsx, and json files."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-report", fmt="all")

    assert set(paths.keys()) >= {"html", "xlsx", "json"}
    for ext, path in paths.items():
        assert path.exists(), f"Output file for '{ext}' not created: {path}"


def test_generate_v2_report_missing_professor_raises_value_error(tmp_path: Path) -> None:
    """generate_v2_report raises ValueError for unknown professor_id."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    with pytest.raises((ValueError, SystemExit)):
        generate_v2_report(project_dir, professor_id="nonexistent-prof", fmt="html")


def test_generate_v2_report_invalid_fmt_raises_value_error(tmp_path: Path) -> None:
    """generate_v2_report raises ValueError for unsupported fmt."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    with pytest.raises(ValueError, match="fmt"):
        generate_v2_report(project_dir, professor_id="prof-report", fmt="pdf")


def test_generate_v2_report_output_path_under_03_report(tmp_path: Path) -> None:
    """Output path is under 03_report/content/v2/ subdirectory."""
    from tube_scout.reporting.content_report import generate_v2_report

    project_dir, _ = _make_project(tmp_path)
    paths = generate_v2_report(project_dir, professor_id="prof-report", fmt="html")

    html_path = paths["html"]
    assert "03_report" in str(html_path)
    assert "v2" in str(html_path)
