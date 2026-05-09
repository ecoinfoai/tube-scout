"""Unit tests for phrase_whitelist service remaining functions (T055 RED).

Tests add_pair_whitelist, add_phrase_whitelist, list_whitelist,
export_whitelist, remove_whitelist, and subtract_phrase_whitelist.
"""

import csv
import sqlite3
from pathlib import Path

import pytest

from tests.fixtures.spec011.fixture_db import build_clean_v2_db
from tube_scout.models.reuse_v2 import MatchSpan, WhitelistPhraseEntry, WhitelistView


def _make_db(tmp_path: Path) -> Path:
    db = build_clean_v2_db(tmp_path / "content_reuse.db")
    conn = sqlite3.connect(str(db))
    conn.execute(
        "INSERT OR IGNORE INTO professor_pool (professor_id, display_name, created_at, created_by) "
        "VALUES ('prof-wl', 'WL Prof', '2026-01-01T00:00:00', 'system')"
    )
    conn.execute(
        "INSERT OR IGNORE INTO comparison_results "
        "(source_video_id, target_video_id, review_status, created_at) "
        "VALUES ('vid_a', 'vid_b', 'UNREVIEWED', '2026-01-01T00:00:00')"
    )
    conn.commit()
    conn.close()
    return db


def _span(text: str, length: float = 10.0) -> MatchSpan:
    return MatchSpan(
        start_a_seconds=0.0, end_a_seconds=length,
        start_b_seconds=0.0, end_b_seconds=length,
        length_seconds=length,
        matched_text_sample=text,
    )


def test_add_pair_whitelist_marks_false_positive(tmp_path: Path) -> None:
    """add_pair_whitelist sets review_status=FALSE_POSITIVE and returns comparison id."""
    from tube_scout.services.phrase_whitelist import add_pair_whitelist

    db = _make_db(tmp_path)
    affected_id = add_pair_whitelist(
        source_video_id="vid_a",
        target_video_id="vid_b",
        reason="test reason",
        db_path=db,
        registered_by="admin",
    )
    assert isinstance(affected_id, int)
    assert affected_id > 0

    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT review_status FROM comparison_results WHERE id = ?", (affected_id,)
    ).fetchone()
    conn.close()
    assert row is not None
    assert row[0] == "FALSE_POSITIVE"


def test_add_pair_whitelist_raises_if_pair_not_found(tmp_path: Path) -> None:
    """add_pair_whitelist raises ValueError if pair not in comparison_results."""
    from tube_scout.services.phrase_whitelist import add_pair_whitelist

    db = _make_db(tmp_path)
    with pytest.raises(ValueError, match="pair not found"):
        add_pair_whitelist(
            source_video_id="nonexistent_a",
            target_video_id="nonexistent_b",
            reason="test",
            db_path=db,
            registered_by="admin",
        )


def test_add_phrase_whitelist_inserts_row(tmp_path: Path) -> None:
    """add_phrase_whitelist inserts a normalized phrase into phrase_whitelist."""
    from tube_scout.services.phrase_whitelist import add_phrase_whitelist

    db = _make_db(tmp_path)
    entry = add_phrase_whitelist(
        professor_id="prof-wl",
        phrase_raw="안녕하세요 여러분",
        reason="habitual greeting",
        db_path=db,
        registered_by="admin",
    )
    assert isinstance(entry, WhitelistPhraseEntry)
    assert entry.professor_id == "prof-wl"
    assert entry.phrase_raw == "안녕하세요 여러분"


def test_add_phrase_whitelist_duplicate_raises_value_error(tmp_path: Path) -> None:
    """Duplicate phrase addition raises ValueError with actionable message."""
    from tube_scout.services.phrase_whitelist import add_phrase_whitelist

    db = _make_db(tmp_path)
    add_phrase_whitelist("prof-wl", "안녕하세요", "first", db, "admin")
    with pytest.raises(ValueError, match="already whitelisted"):
        add_phrase_whitelist("prof-wl", "안녕하세요", "duplicate", db, "admin")


def test_list_whitelist_returns_view(tmp_path: Path) -> None:
    """list_whitelist returns WhitelistView with pair and phrase entries."""
    from tube_scout.services.phrase_whitelist import (
        add_pair_whitelist,
        add_phrase_whitelist,
        list_whitelist,
    )

    db = _make_db(tmp_path)
    add_pair_whitelist("vid_a", "vid_b", "test pair", db, "admin")
    add_phrase_whitelist("prof-wl", "감사합니다", "habitual", db, "admin")

    view = list_whitelist(db)
    assert isinstance(view, WhitelistView)
    assert len(view.pair_entries) >= 1
    assert len(view.phrase_entries) >= 1

    # Filter by professor
    prof_view = list_whitelist(db, professor_id="prof-wl")
    assert all(e.professor_id == "prof-wl" for e in prof_view.phrase_entries)


def test_list_whitelist_kind_filter(tmp_path: Path) -> None:
    """list_whitelist kind='pair' returns only pair entries."""
    from tube_scout.services.phrase_whitelist import (
        add_pair_whitelist,
        add_phrase_whitelist,
        list_whitelist,
    )

    db = _make_db(tmp_path)
    add_pair_whitelist("vid_a", "vid_b", "pair reason", db, "admin")
    add_phrase_whitelist("prof-wl", "오늘 배울", "phrase reason", db, "admin")

    pair_view = list_whitelist(db, kind="pair")
    assert len(pair_view.pair_entries) >= 1
    assert len(pair_view.phrase_entries) == 0

    phrase_view = list_whitelist(db, kind="phrase")
    assert len(phrase_view.pair_entries) == 0
    assert len(phrase_view.phrase_entries) >= 1


def test_export_whitelist_csv(tmp_path: Path) -> None:
    """export_whitelist creates a CSV file and returns its path."""
    from tube_scout.services.phrase_whitelist import add_phrase_whitelist, export_whitelist

    db = _make_db(tmp_path)
    add_phrase_whitelist("prof-wl", "감사합니다", "test", db, "admin")

    output = tmp_path / "whitelist.csv"
    result_path = export_whitelist(db, fmt="csv", output_path=output)
    assert result_path == output
    assert output.exists()
    content = output.read_text(encoding="utf-8")
    assert "감사합니다" in content or "phrase" in content.lower()


def test_export_whitelist_xlsx(tmp_path: Path) -> None:
    """export_whitelist creates an XLSX file with Pair and Phrase sheets."""
    from tube_scout.services.phrase_whitelist import add_phrase_whitelist, export_whitelist

    db = _make_db(tmp_path)
    add_phrase_whitelist("prof-wl", "안녕하세요", "test", db, "admin")

    output = tmp_path / "whitelist.xlsx"
    result_path = export_whitelist(db, fmt="xlsx", output_path=output)
    assert result_path == output
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(str(output))
    assert "Phrase" in wb.sheetnames or "phrase" in [s.lower() for s in wb.sheetnames]


def test_export_whitelist_markdown(tmp_path: Path) -> None:
    """export_whitelist creates a Markdown file."""
    from tube_scout.services.phrase_whitelist import add_phrase_whitelist, export_whitelist

    db = _make_db(tmp_path)
    add_phrase_whitelist("prof-wl", "오늘은", "test", db, "admin")

    output = tmp_path / "whitelist.md"
    result_path = export_whitelist(db, fmt="markdown", output_path=output)
    assert result_path == output
    assert output.exists()
    content = output.read_text(encoding="utf-8")
    assert "|" in content  # markdown table


def test_remove_whitelist_phrase_returns_true(tmp_path: Path) -> None:
    """remove_whitelist removes a phrase entry and returns True."""
    from tube_scout.services.phrase_whitelist import (
        add_phrase_whitelist,
        list_whitelist,
        remove_whitelist,
    )

    db = _make_db(tmp_path)
    entry = add_phrase_whitelist("prof-wl", "제거할 구문", "test", db, "admin")

    # Get the entry id
    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT id FROM phrase_whitelist WHERE professor_id = 'prof-wl'"
    ).fetchone()
    conn.close()
    entry_id = row[0]

    removed = remove_whitelist(db, kind="phrase", entry_id=entry_id)
    assert removed is True

    view = list_whitelist(db, kind="phrase")
    assert all(e.phrase_raw != "제거할 구문" for e in view.phrase_entries)


def test_remove_whitelist_pair_resets_to_unreviewed(tmp_path: Path) -> None:
    """remove_whitelist for kind=pair resets review_status to UNREVIEWED."""
    from tube_scout.services.phrase_whitelist import (
        add_pair_whitelist,
        remove_whitelist,
    )

    db = _make_db(tmp_path)
    affected_id = add_pair_whitelist("vid_a", "vid_b", "test", db, "admin")

    # Get comparison result id
    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT id FROM comparison_results WHERE source_video_id='vid_a'"
    ).fetchone()
    conn.close()
    comparison_id = row[0]

    removed = remove_whitelist(db, kind="pair", entry_id=comparison_id)
    assert removed is True

    conn = sqlite3.connect(str(db))
    status = conn.execute(
        "SELECT review_status FROM comparison_results WHERE id=?", (comparison_id,)
    ).fetchone()[0]
    conn.close()
    assert status == "UNREVIEWED"


def test_subtract_phrase_whitelist_removes_matching_spans(tmp_path: Path) -> None:
    """subtract_phrase_whitelist removes spans matching whitelisted phrases."""
    from tube_scout.services.phrase_whitelist import (
        add_phrase_whitelist,
        subtract_phrase_whitelist,
    )

    db = _make_db(tmp_path)
    add_phrase_whitelist("prof-wl", "안녕하세요", "habitual", db, "admin")

    spans = [
        _span("안녕하세요", 5.0),
        _span("강의 내용 본론", 20.0),
    ]
    remaining, removed_count = subtract_phrase_whitelist("prof-wl", spans, db)
    assert removed_count >= 1
    assert len(remaining) < len(spans)
    texts = [s.matched_text_sample for s in remaining]
    assert "강의 내용 본론" in texts


def test_subtract_phrase_whitelist_returns_int_count(tmp_path: Path) -> None:
    """subtract_phrase_whitelist returns (list[MatchSpan], int) with count >= 0."""
    from tube_scout.services.phrase_whitelist import subtract_phrase_whitelist

    db = _make_db(tmp_path)
    spans = [_span("no match content", 10.0)]
    remaining, count = subtract_phrase_whitelist("prof-wl", spans, db)
    assert isinstance(count, int)
    assert count == 0
    assert len(remaining) == 1
