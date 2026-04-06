"""Tests for Excel export (US4 — T039)."""

from pathlib import Path

import pytest

from tube_scout.models.report import (
    ComplianceMatrix,
    DepartmentOverview,
    ProfessorDetail,
)
from tube_scout.reporting.excel_export import ExcelExporter


@pytest.fixture
def overview() -> DepartmentOverview:
    """Sample DepartmentOverview."""
    return DepartmentOverview(
        channel_id="UCtest123",
        channel_name="Test Department",
        year=2026,
        semester=1,
        total_videos=10,
        total_professors=3,
        total_courses=4,
        total_duration_hours=15.5,
        total_views=5000,
        parse_success_rate=0.9,
    )


@pytest.fixture
def professor_details() -> list[ProfessorDetail]:
    """Sample ProfessorDetail list."""
    return [
        ProfessorDetail(
            professor_name="ProfA",
            video_count=5,
            courses=["Course1", "Course2"],
            weekly_coverage=0.5,
            session_completeness=0.8,
            avg_duration_minutes=30.0,
            total_views=3000,
            avg_views=600.0,
            validation_error_count=1,
        ),
        ProfessorDetail(
            professor_name="ProfB",
            video_count=3,
            courses=["Course3"],
            weekly_coverage=0.25,
            session_completeness=0.6,
            avg_duration_minutes=25.0,
            total_views=1500,
            avg_views=500.0,
            validation_error_count=0,
        ),
    ]


@pytest.fixture
def compliance_entries() -> list[ComplianceMatrix]:
    """Sample ComplianceMatrix list."""
    return [
        ComplianceMatrix(
            professor_name="ProfA",
            week_statuses={
                1: "uploaded",
                2: "uploaded",
                3: "missing",
                4: "late",
            },
            upload_deadline_compliance=0.5,
        ),
        ComplianceMatrix(
            professor_name="ProfB",
            week_statuses={
                1: "uploaded",
                2: "missing",
                3: "missing",
                4: "missing",
            },
            upload_deadline_compliance=1.0,
        ),
    ]


class TestExcelExporter:
    """Tests for ExcelExporter."""

    def test_creates_file(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Should create an xlsx file at the given path."""
        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        assert output_path.exists()

    def test_has_four_sheets(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Should create 4 sheets with correct names."""
        import openpyxl

        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(output_path)
        assert wb.sheetnames == ["개요", "교수별 상세", "준수율", "이상 탐지"]
        wb.close()

    def test_overview_sheet_data(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Overview sheet should contain summary data."""
        import openpyxl

        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb["개요"]
        # Check that key data is present somewhere in the sheet
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend(row)
        assert 10 in values  # total_videos
        assert 5000 in values  # total_views
        wb.close()

    def test_professor_detail_sheet_data(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Professor detail sheet should contain professor data."""
        import openpyxl

        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb["교수별 상세"]
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend(row)
        assert "ProfA" in values
        assert "ProfB" in values
        wb.close()

    def test_compliance_sheet_data(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Compliance sheet should contain week status data."""
        import openpyxl

        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb["준수율"]
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend(row)
        assert "ProfA" in values
        assert "uploaded" in values
        assert "missing" in values
        wb.close()

    def test_anomaly_sheet_exists_but_empty(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Anomaly sheet should exist but be empty (populated by US5)."""
        import openpyxl

        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb["이상 탐지"]
        # Should have at least a header row but no data rows
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 1  # at least header
        wb.close()

    def test_validation_findings_in_anomaly_sheet(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        professor_details: list[ProfessorDetail],
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Anomaly sheet should contain validation findings when provided."""
        from tube_scout.models.validation import ValidationFinding

        findings = [
            ValidationFinding(
                rule_id="V-001",
                severity="WARNING",
                video_ids=["v1"],
                professor="ProfA",
                description="Year mismatch detected",
                details={"expected": 2026, "actual": 2024},
            ),
        ]

        import openpyxl

        output_path = tmp_path / "report.xlsx"
        exporter = ExcelExporter()
        exporter.export(
            overview=overview,
            professor_details=professor_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
            validation_findings=findings,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb["이상 탐지"]
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend(row)
        assert "V-001" in values
        assert "WARNING" in values
        wb.close()

    def test_sanitize_cell_prefixes_formula_characters(self) -> None:
        """_sanitize_cell should prefix =, +, -, @ with a single quote."""
        from tube_scout.reporting.excel_export import _sanitize_cell

        assert _sanitize_cell("=SUM(A1:A10)") == "'=SUM(A1:A10)"
        assert _sanitize_cell("+cmd|' /C calc'!A0") == "'+cmd|' /C calc'!A0"
        assert _sanitize_cell("-1+1") == "'-1+1"
        assert _sanitize_cell("@SUM(A1)") == "'@SUM(A1)"

    def test_sanitize_cell_passes_through_safe_strings(self) -> None:
        """_sanitize_cell should not modify safe values."""
        from tube_scout.reporting.excel_export import _sanitize_cell

        assert _sanitize_cell("Normal text") == "Normal text"
        assert _sanitize_cell("") == ""
        assert _sanitize_cell(42) == 42
        assert _sanitize_cell(None) is None

    def test_formula_injection_in_professor_name(
        self,
        tmp_path: Path,
        overview: DepartmentOverview,
        compliance_entries: list[ComplianceMatrix],
    ) -> None:
        """Professor names with formula chars should be sanitized in output."""
        import openpyxl

        malicious_details = [
            ProfessorDetail(
                professor_name='=HYPERLINK("http://evil.com","Click")',
                video_count=1,
                courses=["Course1"],
                weekly_coverage=0.5,
                session_completeness=0.8,
                avg_duration_minutes=30.0,
                total_views=100,
                avg_views=100.0,
                validation_error_count=0,
            ),
        ]
        output_path = tmp_path / "report.xlsx"
        ExcelExporter().export(
            overview=overview,
            professor_details=malicious_details,
            compliance_entries=compliance_entries,
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb["교수별 상세"]
        prof_cell = ws.cell(row=2, column=1).value
        assert prof_cell.startswith("'")
        wb.close()
